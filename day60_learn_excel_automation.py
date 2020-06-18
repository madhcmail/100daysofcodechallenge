from openpyxl import load_workbook
import pyperclip

wb = load_workbook("compare_id's.xlsx")
ws1 = wb['current_ids']
maxrow = ws1.max_row

listA = []
listB = []
final_list = []

# creating a list for all elements in column A
for col in list('A'):
    for row in range(1,maxrow+1):
        cell = col +str(row)
        listA.append(ws1[cell].value)


# creating a list for all elements in column B
for col in list('B'):
    for row in range(1,maxrow+1):
        cell = col +str(row)
        listB.append(ws1[cell].value)

# comparing the two lists and writing the differnce to a text file
with open("diff.txt", 'w') as f:
    for id in listA:
        if id not in listB:
            f.write(id)
            f.write('\n')


